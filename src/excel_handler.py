from pathlib import Path

import openpyxl


def load_orders(path: str) -> list[dict]:
    """Read all rows from the Excel file and return as a list of dicts.

    The original 1-based row index (including header row) is stored under '_row_index'.
    Rows where product_url is empty are skipped.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    orders = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        record["_row_index"] = row_idx
        # Skip empty rows
        if not record.get("product_url"):
            continue
        orders.append(record)

    wb.close()
    return orders


